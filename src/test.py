import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
import itertools
from PIL import Image, ImageTk
import time
from datetime import datetime
import random
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

class ImageComparisonExperiment:
    def __init__(self):
      self.root = tk.Tk()
      self.root.title("이미지 비교 실험")
      
        # 전체화면 설정
      self.root.attributes('-fullscreen', True)  # 완전 전체화면 모드
      
      # ESC 키를 눌러서 전체화면을 빠져나갈 수 있도록 설정
      self.root.bind('<Escape>', lambda e: self.root.attributes('-fullscreen', False))
      
      # self.stimulus_size_pixels = int((2560 * 10) / 59.8)  # 약 428 픽셀
      # 5cm 크기로 수정 (기존 10cm에서 절반으로)
      
       # 직접 픽셀 크기 설정 (약 189px)
      self.stimulus_size_pixels = 189
      
      # 화면 중앙에 위치
      screen_width = self.root.winfo_screenwidth()
      screen_height = self.root.winfo_screenheight()
      self.root.geometry(f"{2560}x{1440}+{(screen_width-2560)//2}+{(screen_height-1440)//2}")
      
      # 실험 데이터
      self.image_pairs = []
      self.current_pair_index = 0
      self.responses = []
      self.participant_info = {}
      
      # GUI 요소
      self.setup_gui()
      
      # 실험 상태
      self.experiment_started = False
      self.trial_start_time = None
      self.start_time = None
      
    def load_and_resize_image(self, image_path):
      img = Image.open(image_path)
      # 정확히 5cm 크기로 조정
      img = img.resize((self.stimulus_size_pixels, self.stimulus_size_pixels), Image.Resampling.LANCZOS)
      return ImageTk.PhotoImage(img)   
        
    def setup_gui(self):
      # 배경색 설정
      self.root.configure(bg='#7D7D7D')  # 밝은 회색으로 변경
      
      # 중앙 정렬을 위한 메인 프레임
      main_frame = ttk.Frame(self.root)
      main_frame.place(relx=0.5, rely=0.5, anchor='center')
      
      # 제목 레이블
      title_label = ttk.Label(
          main_frame,
          # text="실험 시작",
          text="Figure-Ground Experiment", # 0526 수정사항 1-1
          # text="Depth Perception Experiment" # 0526 수정사항 1-2
          font=('Helvetica', 16, 'bold')
      )
      title_label.pack(pady=(10, 20))
      
      # 참가자 정보 입력 프레임
      self.info_frame = ttk.Frame(main_frame)
      self.info_frame.pack(pady=20)
      
      # 스타일 설정
      style = ttk.Style()
      style.configure('Custom.TLabel', font=('Helvetica', 12))
      style.configure('Custom.TEntry', font=('Helvetica', 12))
      style.configure('Custom.TCombobox', font=('Helvetica', 12))
      
      # 입력 필드 너비 설정
      entry_width = 25
      
      # 참가자 정보 입력 필드 (그리드 레이아웃)
      # 이름 입력
      name_frame = ttk.Frame(self.info_frame)
      name_frame.pack(fill='x', pady=10)
      ttk.Label(
          name_frame,
          text="이름:",
          style='Custom.TLabel',
          width=10
      ).pack(side='left', padx=5)
      self.name_entry = ttk.Entry(
          name_frame,
          width=entry_width,
          font=('Helvetica', 12)
      )
      self.name_entry.pack(side='left', padx=5)
      
      # 나이 입력
      age_frame = ttk.Frame(self.info_frame)
      age_frame.pack(fill='x', pady=10)
      ttk.Label(
          age_frame,
          text="나이:",
          style='Custom.TLabel',
          width=10
      ).pack(side='left', padx=5)
      self.age_entry = ttk.Entry(
          age_frame,
          width=entry_width,
          font=('Helvetica', 12)
      )
      self.age_entry.pack(side='left', padx=5)
      
      # 성별 선택
      gender_frame = ttk.Frame(self.info_frame)
      gender_frame.pack(fill='x', pady=10)
      ttk.Label(
          gender_frame,
          text="성별:",
          style='Custom.TLabel',
          width=10
      ).pack(side='left', padx=5)
      self.gender_var = tk.StringVar()
      gender_combo = ttk.Combobox(
          gender_frame,
          textvariable=self.gender_var,
          values=('남성', '여성'),
          width=entry_width-1,
          style='Custom.TCombobox',
          state='readonly'
      )
      gender_combo.pack(side='left', padx=5)
      
      # 버튼 프레임
      button_frame = ttk.Frame(main_frame)
      button_frame.pack(pady=30)
      
      # 폴더 선택 버튼
      self.folder_button = ttk.Button(
          button_frame,
          text="이미지 폴더 선택",
          command=self.select_folder,
          style='Custom.TButton',
          width=20
      )
      self.folder_button.pack(pady=5)
      
      # 실험 시작 버튼
      self.start_button = ttk.Button(
          button_frame,
          text="실험 시작",
          command=self.start_experiment,
          style='Custom.TButton',
          width=20
      )
      self.start_button.pack(pady=5)
      
      # 버튼 스타일 설정
      style.configure('Custom.TButton', font=('Helvetica', 12))
      
      # 실험 화면 요소들을 담을 프레임 생성
      self.experiment_frame = ttk.Frame(self.root)
      
      # 질문 레이블 생성
      self.question_label = tk.Label(
          self.root,
          # text="다음 이미지 중, 더 가까이 보이는 전경을 선택하세요.",
          text="다음 이미지 중, 전경으로 보이는 것을 선택하세요.", # 0526 수정사항 1-1
          # text="다음 이미지 중, 더 가까이 보이는 것을 선택하세요.", # 0526 수정사항 1-2
          font=('Helvetica', 32),
          fg='#FFFFFF',
          bg='#7D7D7D'
      )
      
      # 이미지 표시 캔버스
      self.canvas = tk.Canvas(
          self.root,
          width=2560,
          height=1440,
          bg='#7D7D7D',
          highlightthickness=0
      )
      
      # 키보드 이벤트 바인딩
      self.root.bind('<Left>', lambda e: self.record_response('left'))
      self.root.bind('<Right>', lambda e: self.record_response('right'))     
        
    def select_folder(self):
        folder_path = filedialog.askdirectory(title="이미지 폴더 선택")
        if folder_path:
            self.load_images(folder_path)
    
    def load_images(self, folder_path):
      '''
      # 이미지 파일 찾기
      image_files = []
      for file in os.listdir(folder_path):
          if file.lower().endswith(('.png', '.jpg', '.jpeg')):
              image_files.append(os.path.join(folder_path, file))
      
      # 모든 가능한 페어 생성
      self.image_pairs = list(itertools.combinations(image_files, 2))
      print(f"총 {len(self.image_pairs)}개의 이미지 페어가 생성되었습니다.")
      '''
      
      # 이미지 파일 찾기
      self.image_files = []  # self.image_pairs 대신 self.image_files 사용
      for file in os.listdir(folder_path):
          if file.lower().endswith(('.png', '.jpg', '.jpeg')):
              self.image_files.append(os.path.join(folder_path, file))
      
      # 이미지 파일 리스트를 무작위로 섞기
      random.shuffle(self.image_files)
      print(f"총 {len(self.image_files)}개의 이미지가 로드되었습니다.")
    
    def start_experiment(self):
      '''
      # 이미지 페어가 로드되었는지 확인
      if not self.image_pairs:
        messagebox.showerror("오류", "먼저 이미지 폴더를 선택해주세요.")
        return
      '''        
      
      # 이미지가 로드되었는지 확인
      if not hasattr(self, 'image_files') or not self.image_files:
          messagebox.showerror("오류", "먼저 이미지 폴더를 선택해주세요.")
          return
      
      # 참가자 정보 검증
      if not self.validate_participant_info():
        return
      
       # 기존 GUI 요소들 제거
      for widget in self.root.winfo_children():
        widget.place_forget()
        widget.pack_forget()
        
      # 실험 프레임 설정 및 표시
      self.experiment_frame.pack(expand=True, fill='both')
      
      # 실험 화면 표시
      self.question_label.pack(pady=(200, 100))      
      self.canvas.pack(expand=True, fill='both')
      
      self.experiment_started = True
      self.current_pair_index = 0
      self.start_time = time.time()      
      
      # 캔버스 업데이트를 위해 약간의 지연 추가
      self.root.after(100, self.show_current_pair)
    
    def show_current_pair(self):
        '''
        if self.current_pair_index >= len(self.image_pairs):
            self.end_experiment()
            return
        '''  
        if self.current_pair_index >= len(self.image_files):
          self.end_experiment()
          return          
        
        # 캔버스 클리어
        self.canvas.delete("all")
        
        # 현재 이미지 페어 로드
        # img1_path, img2_path = self.image_pairs[self.current_pair_index]
        
        # 현재 이미지 로드
        current_image_path = self.image_files[self.current_pair_index]
        
        '''
        # 이미지 크기 조정 및 표시
        img1 = self.load_and_resize_image(img1_path)
        img2 = self.load_and_resize_image(img2_path)
        '''
        
        # 이미지 크기 조정 및 표시
        img = self.load_and_resize_image(current_image_path)
        
        '''
        # 이미지 위치 계산
        canvas_center_x = self.canvas.winfo_width() // 2
        canvas_center_y = self.canvas.winfo_height() // 2
        gap = self.stimulus_size_pixels // 2  # 이미지 사이의 간격
        '''
        
        # 화면의 실제 크기 가져오기
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        
        # 질문 레이블의 높이를 고려한 중앙 좌표 계산
        question_height = 100  # 질문 레이블의 상단 패딩
        spacing = 50         # 질문과 이미지 사이의 간격
        
        # 이미지의 y 좌표를 질문 아래로 조정
        image_y = question_height + spacing + (self.stimulus_size_pixels // 2)
        
        # 이미지 표시 (x 좌표는 화면 중앙, y 좌표는 질문 아래 중앙)
        self.canvas.create_image(canvas_width // 2, image_y,
                        image=img, tags='current_image', anchor='center')
        
        '''
        # 정확한 중앙 좌표 계산
        canvas_center_x = canvas_width // 2
        canvas_center_y = canvas_height // 2
        '''
        
        '''
        # 1.5cm 간격을 픽셀로 변환 (10cm = self.stimulus_size_pixels 픽셀)
        gap = int((self.stimulus_size_pixels * 0.75) / 10) + self.stimulus_size_pixels // 2
        
        # 랜덤하게 이미지 순서 결정
        if random.random() < 0.5:  # 50% 확률로 이미지 순서 변경
            img1, img2 = img2, img1
            img1_path, img2_path = img2_path, img1_path
            self.current_image_order = 'swapped'
        else:
            self.current_image_order = 'normal'
    
        # 이미지 표시
        self.canvas.create_image(canvas_center_x - gap, canvas_center_y,
                              image=img1, tags='img1')
        self.canvas.create_image(canvas_center_x + gap, canvas_center_y,
                              image=img2, tags='img2')                              
        '''                              
        # 이미지 표시
        '''
        self.canvas.create_image(canvas_center_x, canvas_center_y,
                           image=img, tags='current_image', anchor='center')
        '''
        
        '''
        # 이미지 참조 유지
        self.current_images = (img1, img2)
        self.current_image_paths = (img1_path, img2_path)
        '''
        
        # 이미지 참조 유지
        self.current_image = img
        self.current_image_path = current_image_path
        
        # 현재 trial의 시작 시간 기록
        self.trial_start_time = time.time()
    
    def record_response(self, choice):
        # 현재 인덱스가 유효한지 확인
        if self.current_pair_index >= len(self.image_files):
            # 모든 이미지를 다 봤을 때의 처리
            messagebox.showinfo("완료", "모든 이미지 비교가 완료되었습니다.")
            self.root.quit()  # 또는 self.root.destroy()
            return
          
        # 현재 trial의 반응 시간 계산
        response_time = time.time() - self.trial_start_time          

        '''
        # 응답 기록        
        response = {
          'pair_number': self.current_pair_index + 1, 
            'image1': os.path.basename(self.image_pairs[self.current_pair_index][0]),
            'image2': os.path.basename(self.image_pairs[self.current_pair_index][1]),
            'choice': choice,
            'response_time': response_time # 각 trial의 반응 시간으로 수정
        }
        '''
        # 응답 기록
        response = {
          'image_name': os.path.basename(self.current_image_path),
          'choice': choice,
          'response_time': response_time
        }
        self.responses.append(response)
        
        # 다음 이미지 쌍으로 이동
        self.current_pair_index += 1
        
        # 다음 이미지가 있는지 확인
        if self.current_pair_index < len(self.image_files):
            self.show_current_pair()
        else:
            # 모든 이미지를 다 봤을 때의 처리
            self.save_results()  # 응답 저장
            messagebox.showinfo("완료", "모든 이미지 비교가 완료되었습니다.")
            self.root.quit()  # 또는 self.root.destroy()    
    
    def end_experiment(self):
        # 결과 저장
        self.save_results()
        
        # 실험 종료 메시지
        self.canvas.delete("all")
        self.question_label.config(text="실험이 종료되었습니다. 창을 닫아주세요.")
    
    def save_results(self):
      # 저장할 디렉토리 선택
      save_dir = filedialog.askdirectory(title="결과를 저장할 폴더를 선택하세요")
      if not save_dir:  # 사용자가 취소를 누른 경우
          save_dir = "."  # 현재 디렉토리에 저장
      
      '''
      # CSV 파일 생성
      timestamp = datetime.now().strftime('%Y%m%d_%H시_%M분')
      filename = f"results_{self.participant_info['name']}_{timestamp}.csv"
      filepath = os.path.join(save_dir, filename)  # 전체 파일 경로 생성      
      '''
      # 파일명 생성 (더 엄격한 문자 필터링)
      timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')  # 초단위까지 포함
      safe_name = ''.join(c for c in self.participant_info['name'] if c.isalnum() or c in ('-', '_'))
      filename = f"results_{safe_name}_{timestamp}.xlsx"
      filepath = os.path.join(save_dir, filename)
    
      # 기존 파일이 있다면 삭제
      if os.path.exists(filepath):
          try:
              os.remove(filepath)
          except Exception:
              # 파일명 변경
              timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_%f')  # 밀리초 추가
              filename = f"results_{safe_name}_{timestamp}.xlsx"
              filepath = os.path.join(save_dir, filename)
    
      try:
        # 새로운 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "실험결과"

        # 데이터 검증 및 안전한 변환 함수
        def safe_str(value):
            if value is None:
                return ""
            return str(value).replace('\x00', '').strip()

        def safe_number(value):
            try:
                return round(float(value), 3)
            except (ValueError, TypeError):
                return 0.0
          
        # 헤더 정보 작성
        # 헤더 정보 작성 부분을 다음과 같이 수정
        # 제목 셀 병합
        ws.merge_cells('A1:C2')

        # 제목 설정
        title_cell = ws['A1']
        title_cell.value = '실험 결과'
        title_cell.font = Font(size=25, bold=True)  # 폰트 크기 25, 굵게
        title_cell.alignment = Alignment(horizontal='center', vertical='center')  # 가운데 정렬
        
        ws['A3'] = '참가자 이름'
        ws['B3'] = self.participant_info["name"]
        
        ws['A4'] = '나이'
        ws['B4'] = self.participant_info["age"]
        
        ws['A5'] = '성별'
        ws['B5'] = self.participant_info["gender"]
          
        # 데이터 헤더 작성
        headers = ['image_name', 'choice', 'response_time']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
          
         # 데이터 작성 (검증된 데이터만 저장)
        for row_idx, response in enumerate(self.responses, start=9):
            # 이미지 이름 (문자열로 안전하게 변환)
            cell = ws.cell(row=row_idx, column=1)
            cell.value = safe_str(response.get('image_name', ''))
            cell.alignment = Alignment(horizontal='center')
            
            # 선택 (left/right만 허용)
            cell = ws.cell(row=row_idx, column=2)
            choice = safe_str(response.get('choice', ''))
            cell.value = choice if choice in ['left', 'right'] else ''
            cell.alignment = Alignment(horizontal='center')
            
            # 응답 시간 (숫자로 안전하게 변환)
            cell = ws.cell(row=row_idx, column=3)
            cell.value = safe_number(response.get('response_time', 0))
            cell.alignment = Alignment(horizontal='center')
          
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 최대 너비 제한
            ws.column_dimensions[column_letter].width = adjusted_width
        
        '''
        # 파일 저장 전 기존 파일 확인 및 삭제
        if os.path.exists(filepath):
            os.remove(filepath)            
        '''            
          
        # 파일 저장
        wb.save(filepath)                              
        wb.close()
        
        # 저장 완료 메시지
        messagebox.showinfo("저장 완료", f"결과가 다음 위치에 저장되었습니다:\n{filepath}")
          
      except Exception as e:
          messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {str(e)}")
          if 'wb' in locals():
              try:
                  wb.close()
              except:
                  pass   
    
    def validate_participant_info(self):
        name = self.name_entry.get().strip()
        age = self.age_entry.get().strip()
        gender = self.gender_var.get()
        
        if not all([name, age, gender]):
            messagebox.showerror("오류", "모든 참가자 정보를 입력해주세요.")
            return False
        
        try:
            age = int(age)
        except ValueError:
            messagebox.showerror("오류", "나이는 숫자로 입력해주세요.")
            return False
        
        self.participant_info = {
            'name': name,
            'age': age,
            'gender': gender
        }
        return True
    
    def run(self):
        self.root.mainloop()

def main():
    experiment = ImageComparisonExperiment()
    experiment.run()

if __name__ == "__main__":
    main()

